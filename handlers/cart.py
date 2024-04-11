import datetime
import os.path
import time
import pandas
import xlrd
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import xlsxwriter
import getpass
from aiogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from aiogram.utils.callback_data import CallbackData
import config
from services import DataBase
from bot import dp, bot
cb = CallbackData('btn', 'type', 'product_id', 'category_id')
db = DataBase()

async def gen_products(data, user_id):
    keyboard = InlineKeyboardMarkup()
    for i in data:
        count = await db.get_count_in_cart(user_id, i[1])
        count = 0 if not count else sum(j[0] for j in count)
        keyboard.add(InlineKeyboardButton(text=f'{i[2]}: {i[3]}p'))

    keyboard.add(InlineKeyboardButton(text='Назад', callback_data=f'btn:back:-:-'))

    return keyboard

@dp.message_handler(content_types=['document'])
async def files(message: Message):
    try:

        if message.chat.id == config.get_admins(message.chat.id):
            if message.document.file_name == 'Список.xlsx':
                await bot.send_message(message.chat.id,'Подождите несколько секунд, процесс идет')
                await db.del_items_product()
                file_info = await bot.get_file(message.document.file_id)
                downloaded_file = await bot.download_file(file_info.file_path)

                #src = 'C:/Python/Project/tg_bot/files/received/' + message.document.file_name;
                a = os.path.basename(__file__)
                b = os.path.abspath(__file__).replace(a,'')
                src = b+'\\documents\\' + message.document.file_name
                with open(src, 'wb') as new_file:
                    new_file.write(downloaded_file.getvalue())


                book = load_workbook(filename=src)

                sheet = book.active
                count = 1
                while True:
                    if sheet['A'+ str(count)].value == None:
                        break
                    count+=1
                j = 0
                for i in range(2,count):

                    await db.set_product('12345'+str(j),sheet['A'+str(i)].value,sheet['C'+ str(i)].value,sheet['B'+str(i)].value,0)

                    print('прошел')
                    j+=1
                await bot.send_message(message.chat.id, 'Идет изменение категорий подождите еще не много')
                data = await db.get_names()
                for i in data:
                    if str(i[0]).find('Жидко') != -1:
                        await db.update_category_id(1,str(i[0]))
                    if str(i[0]).find('Испар') != -1:
                        await db.update_category_id(2, str(i[0]))
                    if str(i[0]).find('Картр') != -1:
                        await db.update_category_id(3, str(i[0]))
                    if str(i[0]).find('Таб') != -1 or str(i[0]).find('Угл') != -1 or str(i[0]).find('Кал') != -1:
                        await db.update_category_id(4, str(i[0]))
                    if str(i[0]).find('Под') != -1:
                        await db.update_category_id(5, str(i[0]))
                    if str(i[0]).find('Эл') != -1:
                        await db.update_category_id(0, str(i[0]))

                await bot.send_message(message.chat.id,'Данные успешно добавлены')
            if message.document.file_name == 'add.xlsx':
                await bot.send_message(message.chat.id, 'Подождите несколько секунд, процесс идет')
                file_info = await bot.get_file(message.document.file_id)
                downloaded_file = await bot.download_file(file_info.file_path)

                # src = 'C:/Python/Project/tg_bot/files/received/' + message.document.file_name;
                a = os.path.basename(__file__)
                b = os.path.abspath(__file__).replace(a, '')
                src = b + '\\documents\\' + message.document.file_name
                with open(src, 'wb') as new_file:
                    new_file.write(downloaded_file.getvalue())

                book = load_workbook(filename=src)

                sheet = book.active
                count = 1
                while True:
                    if sheet['A' + str(count)].value == None:
                        break
                    count += 1
                j = 0
                product_id = await db.get_product_id_n()
                for j in product_id:
                    prod_id = str(int(j[0] + 1))
                    print(prod_id)
                    for i in range(1, count):
                        await db.set_product(prod_id, sheet['A' + str(i)].value, sheet['C' + str(i)].value,
                                             sheet['B' + str(i)].value, 0)

                        print('прошел')
                        i += 1
                await bot.send_message(message.chat.id, 'Идет изменение категорий подождите еще не много')
                data = await db.get_names()
                for i in data:
                    if str(i[0]).find('Жидко') != -1:
                        await db.update_category_id(1,str(i[0]))
                    if str(i[0]).find('Испар') != -1:
                        await db.update_category_id(2, str(i[0]))
                    if str(i[0]).find('Картр') != -1:
                        await db.update_category_id(3, str(i[0]))
                    if str(i[0]).find('Таб') != -1 or str(i[0]).find('Угл') != -1 or str(i[0]).find('Кал') != -1:
                        await db.update_category_id(4, str(i[0]))
                    if str(i[0]).find('Под') != -1:
                        await db.update_category_id(5, str(i[0]))
                    if str(i[0]).find('Эл') != -1:
                        await db.update_category_id(0, str(i[0]))

                await bot.send_message(message.chat.id,'Данные успешно добавлены')
        else:
            await bot.send_message(message.chat.id,'Увы но наш бот не принимает файлы')
    except Exception as e:
            await bot.send_message(message.chat.id,f'Ошибка: {e}')
@dp.message_handler()
async def shop(message: Message):
    try:
        if message.text == 'Каталог':
            if message.chat.id != config.get_admins(message.chat.id):
                data = await db.get_categories()
                keyboard = InlineKeyboardMarkup()
                for i in data:
                    keyboard.add(InlineKeyboardButton(text=f'{i[0]}', callback_data=f'btn:category:-:{i[1]}'))

                await message.answer('Что хотите купить?', reply_markup=keyboard)
            else:
                data = await db.get_categories()
                keyboard = InlineKeyboardMarkup()
                for i in data:
                    keyboard.add(InlineKeyboardButton(text=f'{i[0]}',callback_data=f'btn:category:-:{i[1]}'))

                await message.answer('Какие товары хочет купить клиент?', reply_markup=keyboard)
        if message.text == 'QR_Code':
            await bot.send_message(message.chat.id,'Ваш qr-code:')
            await bot.send_photo(message.chat.id, photo=open(os.path.abspath('image.jpg'), 'rb'))
        if message.text == 'Помощь':
            await bot.send_message(message.chat.id,'Для обращения напишите в этот чат https://t.me/Nepochatovv')
        if message.text == 'Списание товара':
            if message.chat.id == config.get_admins(message.chat.id):
                data = await db.get_categories()
                keyboard = InlineKeyboardMarkup()
                for i in data:
                    keyboard.add(InlineKeyboardButton(text=f'{i[0]}', callback_data=f'btn:category:-:{i[1]}'))

                await bot.send_message(message.chat.id,'В какой категории товар должен быть списан?',reply_markup=keyboard)
            else:
                bot.send_message(message.chat.id,'Упс.. Вы не туда попали')
        if message.text == 'Добавить товар' and message.chat.id == config.get_admins(message.chat.id):
            data = await db.get_categories()
            keyboard = InlineKeyboardMarkup()
            for i in data:
                keyboard.add(InlineKeyboardButton(text=f'{i[0]}', callback_data=f'btn:categoryadd:-:{i[1]}'))
            await bot.send_message(message.chat.id,'В какой категории нужно добавить товар?',reply_markup=keyboard)

        if message.text == 'Удалить товар' and message.chat.id == config.get_admins(message.chat.id):

            data = await db.get_categories()
            keyboard = InlineKeyboardMarkup()
            for i in data:
                keyboard.add(InlineKeyboardButton(text=f'{i[0]}', callback_data=f'btn:categorydel:-:{i[1]}'))

            await bot.send_message(message.chat.id, 'Нажмите на категорию в который нужно удалить товар',reply_markup=keyboard)
        if message.text == 'Ввод/Вывод учета' and message.chat.id == config.get_admins(message.chat.id):
            keyboard = InlineKeyboardMarkup()
            keyboard.add(InlineKeyboardButton(text=f'Ввод учета',callback_data=f'btn:accouting1:-:-'))
            keyboard.add(InlineKeyboardButton(text=f'Вывод учета', callback_data=f'btn:accouting2:-:-'))
            await bot.send_message(message.chat.id, 'Выбор',reply_markup=keyboard)

        if message.text == 'Завершить работу' and message.chat.id == config.get_admins(message.chat.id):
            name = await db.get_magasin_accout_name(message.chat.id)
            price = await db.get_magasin_accout_price(message.chat.id)
            await bot.send_message(message.chat.id, f'Проданные товары:')
            for i in name:
                await bot.send_message(message.chat.id,f'{i[0]}')
            sums = 0
            for i in price:
                sums += int(i[0])
            await bot.send_message(message.chat.id,f'Выручка: +{sums}руб')
            await bot.send_message(message.chat.id,'Списанные товары:')
            name = await db.get_spis_accout_name(message.chat.id)
            price = await db.get_spis_accout_price(message.chat.id)
            for i in name:
                await bot.send_message(message.chat.id,f'{i[0]}')
            spisan = 0
            for i in price:
                spisan += int(i[0])
            await bot.send_message(message.chat.id, f'Списание: -{spisan}руб')
            await bot.send_message(message.chat.id,f'Итоговая сумма: {sums - spisan}руб')
            await db.drop_tableSpis(message.chat.id)
            await db.drop_tableSum(message.chat.id)
    except Exception as e:
        await bot.send_message(message.chat.id,f'{e}')
        #await bot.send_message(message.chat.id,'Скорее всего еще не наступило 22 часов, или же у вас нет прав')
@dp.callback_query_handler(cb.filter(type='accouting1'))
async def accout(call: CallbackQuery, callback_data: dict):
    await bot.send_message(call.message.chat.id,'Киньте файл в бот и процесс пойдет')

@dp.callback_query_handler(cb.filter(type='accouting2'))
async def accout(call: CallbackQuery, callback_data: dict):
    try:
        data = await db.get_acoproduct()
       # df = pd
        #for i in data:
         #   df += pd.DataFrame({'id':[i[0]],'product_id':[i[1]],'name':[i[2]],'price':[i[3]],'count':[i[4]],'category_id':[i[5]]})
        #df.to_excel()
        #await bot.send_document(call.message.chat.id,open(os.path.abspath('product.xlsx',),'rb'))
        book = xlsxwriter.Workbook('product.xlsx')
        sheet = book.add_worksheet()
        row = 0
        column = 0

        sheet.write(row,column,'id')
        sheet.write(row, column+1, 'product_id')
        sheet.write(row, column+2, 'name')
        sheet.write(row, column+3, 'price')
        sheet.write(row, column+4, 'count')
        sheet.write(row, column+5, 'category_id')
        row = 1
        for i in data:
            sheet.write(row, column, i[0])
            row += 1
        column = 1
        row = 1
        for i in data:
            sheet.write(row, column, i[1])
            row += 1
        column = 2
        row = 1
        for i in data:
            sheet.write(row, column, i[2])
            row += 1
        column = 3
        row = 1
        for i in data:
            sheet.write(row, column, i[3])
            row += 1
        column = 4
        row = 1
        for i in data:
            sheet.write(row, column, i[4])
            row += 1
        column = 5
        row = 1
        for i in data:
            sheet.write(row, column, i[5])
            row += 1
        book.close()
        await bot.send_document(call.message.chat.id,open(os.path.abspath('product.xlsx'),'rb'))
    except Exception as e:
        await bot.send_message(call.message.chat.id,f'{e}')
@dp.callback_query_handler(cb.filter(type='category'))
async def goods(call: CallbackQuery, callback_data: dict):
    try:
        if call.message.chat.id != config.get_admins(call.message.chat.id):
            data = await db.get_products(callback_data.get('category_id'))
            keyboard = InlineKeyboardMarkup()
            for i in data:
                keyboard.add(InlineKeyboardButton(text=f'{i[2]} {i[3]}руб. Осталось {i[4]}',callback_data='gdfgfdgdf'))
            keyboard.add(InlineKeyboardButton(text='Назад', callback_data=f'btn:back:-:-'))
            await call.message.edit_reply_markup(reply_markup=keyboard)
        else:

            data = await db.get_products(callback_data.get('category_id'))
            keyboard = InlineKeyboardMarkup()

            if call.message.text != 'В какой категории товар должен быть списан?':
                for i in data:
                    btnplus = InlineKeyboardButton(text='+',callback_data=f'btn:plus:{i[1]}:{i[5]}')
                    btnminus = InlineKeyboardButton(text='-',callback_data=f'btn:minus:{i[1]}:{i[5]}')
                    btnprice = InlineKeyboardButton(text=f'{i[2]} Осталось {i[4]}',callback_data='gdfgfdgdf')
                    keyboard.row(btnprice)
                    keyboard.row(btnminus,btnplus)
                keyboard.add(InlineKeyboardButton(text='Назад', callback_data=f'btn:back:-:-',))
                await call.message.edit_reply_markup(reply_markup=keyboard)
            else:
                for i in data:
                    btnspis = InlineKeyboardButton(text='Списание',callback_data=f'btn:spis:{i[1]}:{i[5]}')

                    btnprice = InlineKeyboardButton(text=f'{i[2]} Осталось {i[4]}',callback_data='gdfgfdgdf')
                    keyboard.row(btnprice)
                    keyboard.row(btnspis)

                keyboard.add(InlineKeyboardButton(text='Назад', callback_data=f'btn:back:-:-',))
                await call.message.edit_reply_markup(reply_markup=keyboard)
    except Exception as e:
        await bot.send_message(call.message.chat.id,f'{e}')
@dp.callback_query_handler(cb.filter(type='back'))
async def back(call: CallbackQuery):
    try:
        data = await db.get_categories()
        keyboard = InlineKeyboardMarkup()
        for i in data:
            keyboard.add(InlineKeyboardButton(text=f'{i[0]}', callback_data=f'btn:category:-:{i[1]}'))

        await call.message.edit_reply_markup(keyboard)
    except Exception as e:
        await bot.send_message(call.message.chat.id,f'{e}')
@dp.callback_query_handler(cb.filter(type='categorydel'))
async def minus(call: CallbackQuery, callback_data: dict):
    try:
        data = await db.get_products(callback_data.get('category_id'))
        keyboard = InlineKeyboardMarkup()
        for i in data:
            keyboard.add(InlineKeyboardButton(text=f'{i[2]}', callback_data=f'btn:deleteproduct:-:{i[1]}'))
        await call.message.edit_reply_markup(keyboard)
    except Exception as e:
        await bot.send_message(call.message.chat.id,f'{e}')
@dp.callback_query_handler(cb.filter(type='deleteproduct'))
async def minus(call: CallbackQuery, callback_data: dict):
    try:
        product_id = callback_data.get('category_id')

        await db.del_product(product_id)
        await bot.send_message(call.message.chat.id,'Товар удален')
    except Exception as e:
        await bot.send_message(call.message.chat.id,f'{e}')
@dp.callback_query_handler(cb.filter(type='minus'))
async def minus(call: CallbackQuery, callback_data: dict):
    try:
        product_id = callback_data.get('product_id')
        category_id = callback_data.get('category_id')
        count = await db.get_count(product_id)
        print(product_id + ' ' + category_id)

        for i in count:
            minuss = int(i[0]) + 1
            minuses = str(minuss)
            print(minuses)
            result_code = await db.minus_plus(int(product_id), minuses,int(category_id))
            print(result_code)
            price = await db.get_price_product(product_id, category_id)
            for i in price:
                priceres = int(i[0])
                await bot.send_message(call.message.chat.id,
                                       f'Эх, клиент отказался от товара -{priceres}руб из нашей копилки....')

    except Exception as e:
        await bot.send_message(call.message.chat.id,f'{e}')




@dp.callback_query_handler(cb.filter(type='plus'))
async def plus(call: CallbackQuery, callback_data: dict):
    try:
        product_id = callback_data.get('product_id')
        category_id = callback_data.get('category_id')
        count = await db.get_count(product_id)
        print(product_id + ' ' + category_id)

        for i in count:
            pluss = int(i[0]) - 1
            pluses = str(pluss)
            if int(pluses) != 0:
                result_code = await db.minus_plus(int(product_id), pluses, int(category_id))
                price = await db.get_price_product(product_id,category_id)
                for j in price:
                    priceres = int(j[0])
                    data = await db.get_name(product_id)
                    for t in data:
                        names = str(t[0])
                        await db.set_magasin(call.message.chat.id,names,priceres,int(product_id), int(category_id))
                        #await bot.send_message(call.message.chat.id,f'magasin{type(call.message.chat.id)} {type(names)} {type(priceres)} {type(product_id)} {type(category_id)}')
                    await bot.send_message(call.message.chat.id, f'Клиент купил товар, +{priceres}руб в нашу копилочку')
            else:
                await bot.send_message(call.message.chat.id,'Кажется товары закончились, пора бы их пополнить')
    except Exception as e:
        await bot.send_message(call.message.chat.id,f'{e}')

@dp.callback_query_handler(cb.filter(type='spis'))
async def spis(call: CallbackQuery, callback_data: dict):
    try:
        product_id = callback_data.get('product_id')
        category_id = callback_data.get('category_id')
        count = await db.get_count(product_id)
        print(product_id + ' ' + category_id)
        for i in count:
            minuss = int(i[0]) - 1
            minuses = str(minuss)
            if int(minuses) != 0:
                result_code = await db.minus_plus(int(product_id), minuses,int(category_id))
                price = await db.get_price_product(product_id, category_id)
                for j in price:
                    priceres = int(j[0])
                    data = await db.get_name(product_id)
                    for t in data:
                        names = str(t[0])
                        await db.set_magasin(call.message.chat.id, names, priceres, int(product_id), int(category_id))
                        await db.set_spis(call.message.chat.id,names,priceres,product_id,category_id)
                await bot.send_message(call.message.chat.id,'Кажется эта штука не работает,товар занесен в список списанных товаров')
            else:
                bot.send_message(call.message.chat.id,'Кажется товары закончились, пора бы их пополнить')
    except Exception as e:
        await bot.send_message(call.message.chat.id,f'{e}')

@dp.callback_query_handler(cb.filter(type='categoryadd'))
async def spis(call: CallbackQuery, callback_data: dict):
    try:
        await bot.send_message(call.message.chat.id,'Инструкция: для того чтобы добавить товар создайте excel файл add с расширением xlsx, далее  занесите данные в заданном порядке (Название, Остаток, Цена)')
    except Exception as e:
        await bot.send_message(call.message.chat.id,f'{e}')





